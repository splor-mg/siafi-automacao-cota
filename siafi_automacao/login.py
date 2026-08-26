import os
import sys
import glob
import shutil
import subprocess
import time
from datetime import datetime

from dotenv import load_dotenv
from py3270 import Emulator
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fluxo_anular import anular
from fluxo_aprovar import aprovar
from relato import relato

# ---------------------------------------------------------------------------
# Configuracoes
# ---------------------------------------------------------------------------
load_dotenv()
sistema           = os.getenv('SISTEMA')
usuario           = os.getenv('USUARIO')
senha             = os.getenv('SENHA')
unidade_executora = os.getenv('UNIDADE_EXECUTORA')

_onedrive_base = os.getenv('ONEDRIVE_BASE')
siafi_host     = os.getenv('SIAFI_HOST', 'bhmvsb.prodemge.gov.br')
siafi_visivel  = os.getenv('SIAFI_VISIVEL', 'true').lower() == 'true'

month = datetime.today().strftime("%m")

# Marca da linha cuja operacao pode ter chegado ao SIAFI sem o robo ter lido o
# retorno. Precisa ser texto nao vazio: e isso que a tira da fila de pendentes
# e obriga uma conferencia humana antes de reprocessar.
PROGRESSO_INCERTO = 'INTERROMPIDA - VERIFICAR NO SIAFI'

# Pasta de ORIGEM (OneDrive sincronizado) de onde o arquivo a processar e
# MOVIDO para a pasta local. O caminho do Windows  C:/Users/...  e acessado
# a partir do WSL via /mnt/c/...
PASTA_ORIGEM                    = os.path.join(_onedrive_base, 'Robo (IPU 2)', 'Python')
PASTA_DESTINO                   = os.path.join(_onedrive_base, 'Conferencia arquivo robo')
PASTA_REALIZADOS                = os.path.join(_onedrive_base, 'Realizados')
PASTA_REMANEJAMENTOS_REALIZADOS = os.path.join(_onedrive_base, 'Realizados', 'Remanejamentos realizados')

# Pasta local (Linux/WSL) onde o robo realmente atua, para nao depender da
# sincronizacao do OneDrive enquanto grava.
PASTA_LOCAL = os.getenv('PASTA_LOCAL')

# Retry/backoff para operacoes de arquivo dentro de pastas sincronizadas pelo
# OneDrive (podem estar temporariamente travadas por um handle de sync).
MOVER_TENTATIVAS      = int(os.getenv('MOVER_TENTATIVAS', '6'))
MOVER_ESPERA_SEGUNDOS = float(os.getenv('MOVER_ESPERA_SEGUNDOS', '2'))

# Tentativas de conexao com o SIAFI. Antes o laco de conexao era infinito: com
# a VPN fora do ar o robo ficava preso para sempre, sem nunca devolver codigo
# de erro para o robo.bat (a janela do usuario final travava sem explicacao).
CONEXAO_TENTATIVAS      = int(os.getenv('CONEXAO_TENTATIVAS', '10'))
CONEXAO_ESPERA_SEGUNDOS = float(os.getenv('CONEXAO_ESPERA_SEGUNDOS', '3'))


# ---------------------------------------------------------------------------
# Funcoes auxiliares
# ---------------------------------------------------------------------------
def _vazio(v):
    """True para celula vazia (None ou string em branco)."""
    return v is None or (isinstance(v, str) and v.strip() == '')


def _txt_int(v):
    """Converte numero (mesmo vindo como float, ex.: 1451.0) em string inteira."""
    return str(int(float(v)))


def mover(origem, destino, tentativas=MOVER_TENTATIVAS, espera=MOVER_ESPERA_SEGUNDOS):
    """Move 'origem' para 'destino', sobrescrevendo se ja existir.

    Pre-remover o destino evita o erro 'Destination path already exists' do
    shutil.move quando a origem e o destino estao em sistemas de arquivos
    diferentes (caso tipico de WSL local <-> /mnt/c do OneDrive), situacao em
    que o shutil.move faz copy2 + remove em vez de um rename simples.

    Como o destino costuma estar dentro de uma pasta sincronizada pelo
    OneDrive, o arquivo pode estar temporariamente com um handle aberto pelo
    processo de sincronizacao do Windows (PermissionError / WinError 32 via
    WSL). Nesses casos, tenta novamente com backoff antes de desistir."""
    ultimo_erro = None
    for tentativa in range(1, tentativas + 1):
        try:
            if os.path.exists(destino):
                os.remove(destino)
            shutil.move(origem, destino, copy_function=shutil.copyfile)
            return
        except PermissionError as e:
            ultimo_erro = e
            print(
                f"Aviso: destino travado (possivel sincronizacao do OneDrive em andamento). "
                f"Tentativa {tentativa}/{tentativas} de mover '{os.path.basename(origem)}'. "
                f"Aguardando {espera}s..."
            )
            time.sleep(espera)
    print(f"Erro: nao foi possivel mover '{origem}' para '{destino}' apos {tentativas} tentativas.")
    raise ultimo_erro


def encerrar_emulador(em):
    """Encerra o emulador ignorando falhas.

    Usado tanto no caminho feliz quanto no tratamento de erro: se a sessao ja
    caiu, o terminate() pode estourar, e nesse ponto nao ha mais nada a fazer
    alem de garantir que o processo s3270/x3270 nao fique pendurado."""
    if em is None:
        return
    try:
        em.terminate()
    except Exception as e:
        print(f"Aviso: falha ao encerrar o emulador (ignorando): {e}")


def conectar_siafi(host, visivel, tentativas=CONEXAO_TENTATIVAS, espera=CONEXAO_ESPERA_SEGUNDOS):
    """Abre o emulador e conecta ao SIAFI, tentando novamente ate 'tentativas'
    vezes. Devolve o emulador conectado ou aborta com SystemExit(1).

    Alem da recusa de sessao do proprio mainframe ('UNABLE TO ESTABLISH
    SESSION'), trata a falha do Emulator()/connect() em si (VPN desligada,
    x3270 ausente, host inacessivel), que antes subia como traceback cru."""
    ultimo_motivo = None
    for tentativa in range(1, tentativas + 1):
        em = None
        try:
            em = Emulator(visible=visivel)
            em.connect(host)
            em.wait_for_field()
            if not em.string_found(1, 2, 'UNABLE TO ESTABLISH SESSION'):
                return em
            ultimo_motivo = 'o servidor recusou a sessao (UNABLE TO ESTABLISH SESSION)'
        except Exception as e:
            ultimo_motivo = f'{type(e).__name__}: {e}'

        encerrar_emulador(em)
        print(f"Tentativa {tentativa}/{tentativas} de conectar em {host} falhou: {ultimo_motivo}")
        if tentativa < tentativas:
            print(f"Aguardando {espera}s antes de tentar novamente...")
            time.sleep(espera)

    print("")
    # relato (e nao print): acionado pelo Telegram, ninguem esta olhando o
    # console — o motivo real da falha precisa chegar ao grupo.
    relato('erro',
           f"Nao foi possivel conectar ao SIAFI apos {tentativas} tentativas.\n"
           f"Ultimo motivo: {ultimo_motivo}\n"
           "Verifique se a VPN esta conectada e se o SIAFI esta no ar, e rode novamente.")
    raise SystemExit(1)


def resgatar_planilha(wb, ws, caminho_local, caminho_destino):
    """Devolve a planilha para a pasta de conferencia depois de uma falha.

    Sem isso, qualquer erro depois do passo 1 deixava o arquivo preso na pasta
    local do WSL: a pasta de origem ja estava vazia e os remanejamentos ja
    tinham sido movidos para Realizados, entao a execucao seguinte nao
    encontrava a planilha em lugar nenhum e so restava recuperacao manual.

    Como cada linha processada ja tem a coluna Progresso gravada e salva na
    hora, o arquivo devolvido e retomado exatamente de onde parou na proxima
    execucao (o consolida.py o encontra na pasta de conferencia e o login.py
    so processa as linhas com Progresso vazio).

    Esta funcao NUNCA propaga excecao: ela roda durante o tratamento de outro
    erro, e deixar uma falha de resgate mascarar o erro original so dificulta
    o diagnostico."""
    if not os.path.exists(caminho_local):
        return

    try:
        formatar_planilha(ws)
        wb.save(caminho_local)
    except Exception as e:
        print(f"Aviso: nao foi possivel formatar/salvar a planilha no resgate: {e}")

    try:
        mover(caminho_local, caminho_destino)
        print(f"Planilha com o progresso parcial devolvida para: {caminho_destino}")
    except Exception as e:
        print("")
        print(f"ERRO: a planilha ficou parada em '{caminho_local}' e nao pode ser devolvida: {e}")
        print(f"Mova o arquivo manualmente para '{caminho_destino}' antes da proxima execucao.")


def organizar_realizados(pasta_origem, pasta_destino):
    """Move os .xlsx soltos em 'pasta_origem' para 'pasta_destino'.
    Se ja existir um arquivo com o mesmo nome, adiciona sufixo (1), (2), etc."""
    os.makedirs(pasta_destino, exist_ok=True)
    arquivos = [
        f for f in os.listdir(pasta_origem)
        if f.endswith('.xlsx') and os.path.isfile(os.path.join(pasta_origem, f))
    ]
    for nome in arquivos:
        origem = os.path.join(pasta_origem, nome)
        destino = os.path.join(pasta_destino, nome)
        if os.path.exists(destino):
            base, ext = os.path.splitext(nome)
            contador = 1
            while os.path.exists(destino):
                destino = os.path.join(pasta_destino, f"{base} ({contador}){ext}")
                contador += 1
        shutil.move(origem, destino, copy_function=shutil.copyfile)
        print(f"Organizando: {nome} -> {os.path.basename(destino)}")


def localizar_arquivo(pasta):
    """Retorna o caminho do .xlsx mais recente da pasta, ignorando arquivos
    temporarios de lock do Excel (que comecam com '~$')."""
    candidatos = [
        c for c in glob.glob(os.path.join(pasta, '*.xlsx'))
        if not os.path.basename(c).startswith('~$')
    ]
    if not candidatos:
        raise FileNotFoundError(f"Nenhum arquivo .xlsx encontrado em: {pasta}")
    candidatos.sort(key=os.path.getmtime, reverse=True)  # mais recente primeiro
    if len(candidatos) > 1:
        print("Aviso: ha mais de um .xlsx na pasta. Usando o mais recente:")
        print(f"   -> {os.path.basename(candidatos[0])}")
    return candidatos[0]


def localizar_aba(wb):
    """Localiza a aba que contem os dados (a que tem as colunas 'Progresso' e
    'UO_COD' no cabecalho). Assim o script funciona independente do nome da
    aba ('Planilha1', 'Remanejamento Cota Orcamentaria', etc.)."""
    for ws in wb.worksheets:
        cabec = [ws.cell(row=1, column=c).value for c in range(1, (ws.max_column or 0) + 1)]
        if 'Progresso' in cabec and 'UO_COD' in cabec:
            return ws
    return wb.active


def traduzir_progresso(retorno):
    """Converte a mensagem crua do SIAFI no texto que vai para a coluna
    'Progresso'. Mensagens conhecidas viram um texto amigavel; qualquer outro
    retorno e tratado como sucesso ('Ok').
    Para incluir novas mensagens, basta acrescentar uma linha no mapa abaixo."""
    if retorno is None:
        return 'Ok'
    retorno = retorno.strip()

    if retorno.startswith("E90 - SALDO ZERADO NA CONTA"):
        return 'Saldo zerado na conta'

    mapa = {
        "0139- VALOR A APROVAR MAIOR QUE SALDO DISPONIVEL NO PROJ/ATIV.":
            'Valor a aprovar maior que o saldo disponível',
        "Inconsistencia no Registro da Contabilizacao":
            'Erro de Saldo Contábil',
        "0139- PROGRAMA DE TRABALHO NAO ENCONTRADO PARA GM/FP.":
            'Programa de trabalho não encontrado para GM/FP',
        "0139- VALORES A ANULAR MAIOR QUE SALDO DISPONIVEL.":
            'Valor a anular maior que o saldo disponível',
        "0139- PROJ/ATIV OU FONTE/PROC./IAG INEXISTENTE PARA UO":
            'Proj/Ativ ou Fonte/Proc./IAG inexistente para a UO',
        "0101- GRUPO DESPESA INEXISTENTE(S).":
            'Grupo de despesa inexistente',
        "0139- ELEMENTO/ITEM NAO MARCADO PARA UO BENEFICIADA.":
            'Elemento/item não marcado para a UO beneficiada',
        "SALDO DE CREDITO ORCAMENTARIO A APROVAR POR PROJ/ATIV ZERADO.":
            'Saldo de crédito a aprovar zerado',
        # Unico item do mapa que nao vem do mainframe: e o retorno sintetico
        # que login.py atribui a uma linha sem valor de anulacao/aprovacao,
        # que por isso nunca chega a ser enviada ao SIAFI.
        "Linha sem valor de anulação/aprovação":
            'Linha sem valor de anulação/aprovação',
    }
    return mapa.get(retorno, 'Ok')


def formatar_planilha(ws):
    """Aplica formatacao visual gerencial na aba: cabecalho colorido, valores
    numericos formatados, zebra nas linhas, coluna Progresso com cor condicional
    e larguras ajustadas ao conteudo."""
    AZUL_ESCURO  = PatternFill('solid', fgColor='1F4E79')
    AZUL_CLARO   = PatternFill('solid', fgColor='D6E4F0')
    BRANCO       = PatternFill('solid', fgColor='FFFFFF')
    VERDE        = PatternFill('solid', fgColor='C6EFCE')
    AMARELO      = PatternFill('solid', fgColor='FFEB9C')
    VERMELHO     = PatternFill('solid', fgColor='FFC7CE')
    FONTE_BRANCA = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    FONTE_NORMAL = Font(name='Calibri', size=10)
    BORDA = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )
    COLS_VALOR = {'Anular', 'Aprovar'}

    max_col = ws.max_column
    max_row = ws.max_row

    # Mapeia nome da coluna -> indice
    cabec = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}

    # Remove linhas de grade
    ws.sheet_view.showGridLines = False

    # Cabecalho
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill   = AZUL_ESCURO
        cell.font   = FONTE_BRANCA
        cell.border = BORDA
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    ws.row_dimensions[1].height = 20

    # Linhas de dados
    for r in range(2, max_row + 1):
        zebra = AZUL_CLARO if r % 2 == 0 else BRANCO
        for c in range(1, max_col + 1):
            cell      = ws.cell(row=r, column=c)
            cell.fill = zebra
            cell.font = FONTE_NORMAL
            cell.border = BORDA

            col_nome = ws.cell(row=1, column=c).value

            # Formata colunas de valor monetario (alinha direita)
            if col_nome in COLS_VALOR and cell.value not in (None, ''):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Cor condicional na coluna Progresso
        if 'Progresso' in cabec:
            prog_cell = ws.cell(row=r, column=cabec['Progresso'])
            valor = (prog_cell.value or '').strip()
            if valor == 'Ok':
                prog_cell.fill = VERDE
            elif valor != '':
                prog_cell.fill = VERMELHO if 'zerado' in valor.lower() or 'maior' in valor.lower() or 'inexistente' in valor.lower() else AMARELO

    # Congela a primeira linha
    ws.freeze_panes = 'A2'

    # Ajusta largura: usa o maior entre o titulo do cabecalho e o conteudo
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)


def montar_data_row(get, month):
    """Monta o dicionario data_row a partir de uma linha da planilha.
    'get' e uma funcao que recebe o nome da coluna e devolve o valor da celula."""
    dr = {'month': month}
    dr['uo']          = _txt_int(get('UO_COD'))
    dr['grupo']       = _txt_int(get('Grupo'))
    dr['iag']         = _txt_int(get('IAG'))
    dr['fonte']       = _txt_int(get('Fonte'))
    dr['procedencia'] = _txt_int(get('IPU'))
    dr['acao']        = _txt_int(get('Ação'))

    g = get('GLOBAL')
    dr['tipo_global'] = g.strip().lower() if (isinstance(g, str) and g.strip() != '') else '0'

    am = get('AMARRADO')
    if not _vazio(am):
        amarrado = _txt_int(am).zfill(4)   # garante 4 digitos (ex.: 308 -> '0308')
        dr['tipo_amarrado'] = amarrado
        dr['elemento'] = amarrado[:2]      # dois primeiros digitos
        dr['item']     = amarrado[2:]      # dois ultimos digitos
    else:
        dr['tipo_amarrado'] = '0'
        dr['elemento'] = '0'
        dr['item']     = '0'

    uof = get('UO Financiadora')
    dr['uo_financiadora'] = _txt_int(uof) if not _vazio(uof) else '0'

    av = get('Anular')
    pv = get('Aprovar')
    dr['valor_anulacao']  = int(round(float(av) * 100)) if not _vazio(av) else 0
    dr['valor_aprovacao'] = int(round(float(pv) * 100)) if not _vazio(pv) else 0

    # valor a preencher: usa anulacao se houver, senao aprovacao
    dr['valor'] = dr['valor_anulacao'] if dr['valor_anulacao'] != 0 else dr['valor_aprovacao']
    return dr


# ===========================================================================
# Execucao
# ===========================================================================
if __name__ == "__main__":

    # -----------------------------------------------------------------------
    # 0) Executa consolida.py e segue direto para o fluxo no SIAFI.
    # -----------------------------------------------------------------------
    script_consolida = os.path.join(os.path.dirname(__file__), 'consolida.py')
    print("Executando consolida.py...")
    consolida = subprocess.run([sys.executable, script_consolida])
    if consolida.returncode != 0:
        # Sem check=True: o CalledProcessError subia como traceback cru, e o
        # grupo recebia 'FALHOU (codigo 1)' sem motivo nenhum. O consolida.py
        # ja emitiu o relato especifico; aqui so encerramos limpo, dizendo onde
        # esta o detalhe.
        relato('erro', 'A consolidação das planilhas não passou. Nenhuma linha '
                       'foi processada. Use /log para o relatório completo.')
        raise SystemExit(consolida.returncode)

    # -----------------------------------------------------------------------
    # 1) Move o arquivo da pasta de origem para a pasta local e abre a copia
    # -----------------------------------------------------------------------
    os.makedirs(PASTA_LOCAL, exist_ok=True)
    os.makedirs(PASTA_DESTINO, exist_ok=True)

    # Pasta vazia nao e falha, e o caso comum de acionar sem ter posto
    # planilha nenhuma. Sem este tratamento o FileNotFoundError subia como
    # traceback e o grupo recebia 'FALHOU (codigo 1)' sem explicacao alguma.
    try:
        arquivo_origem = localizar_arquivo(PASTA_ORIGEM)
    except FileNotFoundError:
        relato('aviso', 'Nenhuma planilha para processar. Coloque o arquivo na '
                        'pasta Remanejamentos e acione de novo.')
        raise SystemExit(0)

    nome_arquivo    = os.path.basename(arquivo_origem)
    caminho_local   = os.path.join(PASTA_LOCAL, nome_arquivo)
    caminho_destino = os.path.join(PASTA_DESTINO, nome_arquivo)

    # A partir daqui o arquivo NAO existe mais na pasta de origem.
    mover(arquivo_origem, caminho_local)
    print(f"Arquivo movido da pasta de origem para a pasta local: {caminho_local}")

    wb = load_workbook(caminho_local)
    ws = localizar_aba(wb)
    col = {ws.cell(row=1, column=c).value: c
           for c in range(1, ws.max_column + 1)
           if ws.cell(row=1, column=c).value}

    # -----------------------------------------------------------------------
    # 2) Identifica as linhas pendentes (tem dados, mas a coluna Progresso
    #    ainda esta vazia). E o caso da execucao mais recente.
    # -----------------------------------------------------------------------
    pendentes = [
        r for r in range(2, ws.max_row + 1)
        if not _vazio(ws.cell(row=r, column=col['UO_COD']).value)
        and _vazio(ws.cell(row=r, column=col['Progresso']).value)
    ]

    if not pendentes:
        relato('aviso', 'Nenhuma linha pendente: todas ja tem Progresso '
                        'preenchido na planilha.')
        mover(caminho_local, caminho_destino)
        print(f"Arquivo movido para a pasta de conferencia: {caminho_destino}")
        raise SystemExit(0)

    relato('pendentes',
           f"{len(pendentes)} linha(s) pendente(s): "
           f"{', '.join(str(p) for p in pendentes)}",
           linhas=pendentes)

    # -----------------------------------------------------------------------
    # 3 a 6) Daqui em diante o arquivo ja saiu da pasta de origem e o SIAFI
    #        sera aberto. Tudo fica dentro de um try/except/finally para
    #        garantir duas coisas que antes so aconteciam no caminho feliz:
    #          - o emulador SEMPRE e encerrado, sem deixar processo s3270/
    #            x3270 pendurado nem sessao aberta no mainframe;
    #          - a planilha SEMPRE volta para a pasta de conferencia com o
    #            progresso parcial, em vez de ficar orfa na pasta local do
    #            WSL (de onde nenhuma execucao seguinte conseguia recupera-la).
    # -----------------------------------------------------------------------
    em = None
    # Linha cuja operacao pode ter chegado ao SIAFI sem o retorno ter sido lido.
    linha_em_curso = None
    try:
        # -------------------------------------------------------------------
        # 3) Login no SIAFI
        # -------------------------------------------------------------------
        em = conectar_siafi(siafi_host, siafi_visivel)

        em.fill_field(19, 13, sistema, 8)
        em.fill_field(20, 13, usuario, 8)
        em.fill_field(21, 13, senha, 8)
        em.send_enter()

        max_tentativas = 10
        tentativas = 0
        while tentativas < max_tentativas:
            time.sleep(1)
            try:
                em.send_enter()
                if em.string_found(1, 13, 'Logon executado com sucesso'):
                    relato('login', "Login no SIAFI realizado")
                    break
                else:
                    print(f"Tentativa {tentativas + 1} - tela intermediária, avançando...")
                    em.send_enter()
            except Exception:
                # Exception (e nao bare except) para nao engolir um Ctrl+C do
                # usuario, que antes virava "tela de aviso detectada".
                print(f"Tentativa {tentativas + 1} - tela de aviso detectada, passando...")
                em.send_enter()
            tentativas += 1

        if tentativas == max_tentativas:
            relato('erro', "Não foi possível fazer login após várias tentativas.")
            raise SystemExit(1)

        em.fill_field(1, 2, sistema, 4)
        em.send_enter()

        # nova tela buscando login...
        max_tentativas = 10
        tentativas = 0
        while tentativas < max_tentativas:
            time.sleep(1)
            try:
                em.send_enter()
                if em.string_found(22, 11, 'Unidade Executora'):
                    print("Texto encontrado")
                    break
                else:
                    print(f"Tentativa {tentativas + 1} - tela intermediária, avançando...")
                    em.send_enter()
            except Exception:
                print(f"Tentativa {tentativas + 1} - tela de aviso detectada, passando...")
                em.send_enter()
            tentativas += 1

        if tentativas == max_tentativas:
            relato('erro', "Não foi possível fazer login após várias tentativas.")
            raise SystemExit(1)

        # Entrar com a Unidade Executora
        em.fill_field(22, 30, unidade_executora, 7)
        em.send_enter()
        em.wait_for_field()
        # Fim do login

        # Entrar em 03 - Movimentacao Orcamentaria
        em.fill_field(21, 19, '03', 2)
        em.send_enter()
        em.wait_for_field()

        # Entrar em 02 - Aprovacao de Cota Orcamentaria
        em.fill_field(21, 19, '02', 2)
        em.send_enter()
        em.wait_for_field()

        # -------------------------------------------------------------------
        # 4) Processa cada linha pendente e grava o resultado na coluna
        #    Progresso
        # -------------------------------------------------------------------
        for r in pendentes:
            get = lambda nome: ws.cell(row=r, column=col[nome]).value
            data_row = montar_data_row(get, month)

            # Remanejamentos so sao permitidos para IAG 0.
            if data_row['iag'] == '1':
                relato('linha_pulada', f"Linha {r}: IAG 1, pulando.",
                       linha=r, motivo='IAG 1')
                ws.cell(row=r, column=col['Progresso']).value = 'IAG 1 - Não Realizado'
                wb.save(caminho_local)
                continue

            # Linha sem GLOBAL e sem AMARRADO nao e processavel no SIAFI:
            # registra o motivo e segue para a proxima.
            if data_row['tipo_global'] != 'x' and data_row['tipo_amarrado'] == '0':
                relato('linha_pulada',
                       f"Linha {r}: sem GLOBAL/AMARRADO definido, pulando.",
                       linha=r, motivo='sem GLOBAL/AMARRADO')
                ws.cell(row=r, column=col['Progresso']).value = 'Linha sem GLOBAL/AMARRADO definido'
                wb.save(caminho_local)
                continue

            if data_row['valor_anulacao'] != 0:
                operacao = 'anulação'
            elif data_row['valor_aprovacao'] != 0:
                operacao = 'aprovação'
            else:
                operacao = 'sem valor'

            # A frase "realizando procedimento de X" so faz sentido quando ha
            # de fato um procedimento a realizar no SIAFI.
            prefixo = ('' if operacao == 'sem valor'
                       else f"realizando procedimento de {operacao}\n")

            relato('linha',
                   f"{prefixo}"
                   f"Processando linha {r} | UO: {data_row['uo']}, "
                   f"Grupo: {data_row['grupo']}, Acao: {data_row['acao']}, "
                   f"Fonte: {data_row['fonte']}, "
                   f"Procedencia: {data_row['procedencia']}, "
                   f"Valor: {data_row['valor']}",
                   linha=r, operacao=operacao, uo=data_row['uo'],
                   acao=data_row['acao'], fonte=data_row['fonte'],
                   valor=data_row['valor'])

            # A partir daqui o SIAFI pode ser tocado. Se algo falhar antes de
            # o Progresso ser gravado, esta linha fica em estado incerto.
            linha_em_curso = r

            retorno = None
            if data_row['valor_anulacao'] != 0:
                retorno = anular(em, data_row)
            elif data_row['valor_aprovacao'] != 0:
                retorno = aprovar(em, data_row)
            else:
                retorno = 'Linha sem valor de anulação/aprovação'

            # Grava o resultado e salva imediatamente (resiliencia: se o SIAFI
            # travar no meio, o progresso ja concluido fica registrado).
            progresso = traduzir_progresso(retorno)
            relato('resultado', f"Linha {r}: {progresso}",
                   linha=r, ok=(progresso == 'Ok'), progresso=progresso)
            ws.cell(row=r, column=col['Progresso']).value = progresso
            wb.save(caminho_local)
            linha_em_curso = None

        relato('fim', 'Fluxo finalizado')
        # Encerra o emulador ja aqui (como antes) para a janela do x3270 fechar
        # assim que o SIAFI nao e mais necessario. O finally cobre os casos de
        # erro, e o em = None evita encerrar duas vezes.
        encerrar_emulador(em)
        em = None

        # -------------------------------------------------------------------
        # 5) Formata e move o arquivo atualizado para a pasta de conferencia.
        #    Este e o ponto original do erro: destino dentro do OneDrive pode
        #    estar com lock de sincronizacao. mover() agora tenta novamente com
        #    backoff antes de propagar a excecao.
        # -------------------------------------------------------------------
        formatar_planilha(ws)
        wb.save(caminho_local)
        mover(caminho_local, caminho_destino)
        relato('planilha_final',
               f"Planilha atualizada e movida para a pasta de conferencia: {caminho_destino}",
               arquivo=os.path.basename(caminho_destino))

        # -------------------------------------------------------------------
        # 6) Organiza os .xlsx soltos em Realizados -> Remanejamentos realizados.
        # -------------------------------------------------------------------
        organizar_realizados(PASTA_REALIZADOS, PASTA_REMANEJAMENTOS_REALIZADOS)
        print("Pasta Realizados organizada.")

    except BaseException as e:
        # BaseException (nao Exception) para cobrir tambem o Ctrl+C do usuario
        # e os SystemExit(1) de falha de login: em todos esses casos a planilha
        # ja esta na pasta local e precisa voltar para a de conferencia.
        print("")
        if isinstance(e, SystemExit):
            relato('erro', "Execução interrompida antes de concluir todas as linhas.")
        else:
            relato('erro', f"Execução interrompida por erro: {type(e).__name__}: {e}")

        # A linha que estava em curso pode ter chegado ao SIAFI sem o retorno
        # ter sido lido. Deixar o Progresso vazio faria a proxima execucao
        # reprocessa-la as cegas — e duplicar a operacao, se ela tiver entrado.
        # Marcada, ela sai da fila e alguem precisa conferir antes de liberar.
        if linha_em_curso is not None:
            ws.cell(row=linha_em_curso, column=col['Progresso']).value = PROGRESSO_INCERTO
            relato('erro',
                   f"ATENÇÃO: a linha {linha_em_curso} parou no meio. Não dá para "
                   f"saber se a operação entrou no SIAFI. Confira lá antes de "
                   f"rodar de novo — ela NÃO será reprocessada automaticamente. "
                   f"Para liberá-la, apague a coluna Progresso dessa linha.")

        resgatar_planilha(wb, ws, caminho_local, caminho_destino)
        raise

    finally:
        encerrar_emulador(em)